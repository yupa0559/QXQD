"""
外观不良智能管理系统 V3 - Python/Streamlit 版
修复：HTML源码渲染、隐藏Streamlit水印/菜单、自动同步、全中文
"""

import streamlit as st
import json
import time
import requests
import io
import base64
from datetime import date, datetime
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════
# CONFIG
# ════════════════════════════════════════════════
DEEPSEEK_KEY = "sk-c285636d450b468ba3b437891743d3d5"
DS_API       = "https://api.deepseek.com/v1/chat/completions"

SB_URL   = "https://scgscpfjpdhtqdzkszty.supabase.co"
SB_KEY   = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InNjZ3NjcGZqcGRodHFkemtzenR5Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzU4MjkxMzEsImV4cCI6MjA5MTQwNTEzMX0.Yl93_IKvqfL50GJduPfaK-oijnSN5zZCwXYifMO_4U0"
SB_TABLE = "defect_records"

SEV_LABEL = {"low":"低·轻微","medium":"中·一般","high":"高·严重","critical":"致命·报废"}
SEV_COLOR = {
    "low":      ("#107c10","#e6f4ea"),
    "medium":   ("#c55a11","#fff4e5"),
    "high":     ("#c72a2a","#fde8e8"),
    "critical": ("#7b5ea7","#f0e8fa"),
}
SHIFT_LABEL = {"":"—","day":"☀ 白班","night":"🌙 夜班"}

DEFAULT_TYPES = ["缺料","缩水","批锋","气纹","拉伤","划痕","氧化","开裂","变形","其它"]
DEFAULT_MOLDS = [
    {"id":"M24-001","name":"M24-001","partNo":"","partVer":"","productName":""},
    {"id":"M24-002","name":"M24-002","partNo":"","partVer":"","productName":""},
]

# ════════════════════════════════════════════════
# 页面配置（必须是第一个 st 调用）
# ════════════════════════════════════════════════
st.set_page_config(
    page_title="外观不良智能管理系统 V3",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={}
)

# 全局 CSS：统一视觉风格（中文界面）
st.markdown("""
<style>
/* ── 主题变量 ── */
:root{
    --bg-top:#f5f8fc;
    --bg-bottom:#eef3f9;
    --panel:#ffffff;
    --line:#d7e0ec;
    --text:#1f2a3d;
    --muted:#6d7f99;
    --brand:#0f6b9f;
    --brand-2:#0f8b7f;
}

/* ── 隐藏多余默认 UI ── */
#MainMenu                              { visibility: hidden !important; }
header[data-testid="stHeader"]         { display: none !important; }
[data-testid="stToolbar"]              { display: none !important; }
footer                                 { visibility: hidden !important; }
[data-testid="stDecoration"]           { display: none !important; }
.viewerBadge_container__r5tak         { display: none !important; }
a[href*="streamlit.io"]                { display: none !important; }

/* ── 布局 ── */
body, .stApp {
    background: linear-gradient(180deg, var(--bg-top) 0%, var(--bg-bottom) 100%) !important;
    color: var(--text) !important;
}
.block-container {
    padding-top: 1.1rem !important;
    padding-bottom: 1.25rem !important;
    max-width: 1420px !important;
}
[data-testid="stSidebar"] {
    min-width: 300px !important;
    max-width: 320px !important;
}
[data-testid="stSidebar"] > div:first-child {
    background: linear-gradient(180deg,#f7fbff,#f0f6fb) !important;
    border-right: 1px solid var(--line) !important;
}

/* ── 字体 ── */
html,body,[class*="css"] { font-family:"Noto Sans SC","Microsoft YaHei",sans-serif !important; }

/* ── 指标卡 ── */
[data-testid="metric-container"] {
    background: var(--panel);
    border: 1px solid var(--line);
    border-radius: 12px;
    padding: 12px 16px;
    box-shadow: 0 4px 14px rgba(25, 61, 107, .07);
}
[data-testid="stMetricLabel"] { font-size:12px !important; color:var(--muted) !important; }
[data-testid="stMetricValue"] { font-size:28px !important; font-weight:900 !important; }

/* ── Tab ── */
[data-testid="stTabs"] [role="tab"]{
    border-radius: 10px !important;
    border: 1px solid #d9e3ef !important;
    background: #fff !important;
    margin-right: 6px !important;
    padding: 8px 14px !important;
    font-size: 13px !important;
    font-weight: 700 !important;
}
[data-testid="stTabs"] [aria-selected="true"]{
    border-color: #8bb9db !important;
    color: #0d4670 !important;
    box-shadow: 0 1px 8px rgba(42,90,137,.10) !important;
}

/* ── 按钮/输入框 ── */
.stButton>button {
    border-radius: 9px;
    font-size: 13px;
    font-weight: 700;
    border: 1px solid #d2ddeb;
    min-height: 40px;
}
.stButton>button[kind="primary"]{
    background: linear-gradient(135deg, var(--brand), var(--brand-2)) !important;
    color: #fff !important;
    border: 0 !important;
}
.stTextInput input,.stTextArea textarea, .stSelectbox [data-baseweb="select"] > div {
    font-size: 13px !important;
    border-radius: 9px !important;
}

/* ── 表格 ── */
.stDataFrame { font-size:12px; }

/* ── 表单边框 ── */
[data-testid="stForm"] {
    border:1px solid var(--line) !important;
    border-radius:12px !important;
    padding:12px !important;
    background: #f9fbfd;
}

/* 分割线和提示风格 */
hr {
    border-top: 1px solid #d6e1ee !important;
}
[data-testid="stAlert"]{
    border-radius: 10px !important;
    border: 1px solid #d6e1ee !important;
}
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════
# SESSION STATE
# ════════════════════════════════════════════════
def init_state():
    defaults = {
        "data_store":       [],
        "mold_list":        DEFAULT_MOLDS.copy(),
        "defect_types":     DEFAULT_TYPES.copy(),
        "saved_checklists": [],
        "sync_status":      ("ing","连接中…"),
        "edit_record":      None,
        "synced_once":      False,
        "last_report":      "",
        "last_report_range":"",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def normalize_mold_code(raw: str) -> str:
    return (raw or "").strip().upper()


# ════════════════════════════════════════════════
# SUPABASE REST
# ════════════════════════════════════════════════
def _hdrs(merge=False):
    h = {"apikey":SB_KEY,"Authorization":f"Bearer {SB_KEY}","Content-Type":"application/json"}
    if merge: h["Prefer"] = "resolution=merge-duplicates,return=minimal"
    return h

def sb_get_all():
    try:
        r = requests.get(f"{SB_URL}/rest/v1/{SB_TABLE}?select=*&order=id.desc",
                         headers=_hdrs(), timeout=12)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        st.session_state.sync_status = ("err",f"拉取失败：{e}")
        return []

def sb_upsert(record):
    row = local_to_cloud(record)
    try:
        r = requests.post(f"{SB_URL}/rest/v1/{SB_TABLE}",
                          headers=_hdrs(merge=True), json=row, timeout=12)
        r.raise_for_status()
        return True
    except Exception as e:
        st.session_state.sync_status = ("err",f"上传失败：{e}")
        return False

def sb_delete(rid):
    try:
        r = requests.delete(f"{SB_URL}/rest/v1/{SB_TABLE}?id=eq.{rid}",
                            headers=_hdrs(), timeout=12)
        r.raise_for_status()
        return True
    except Exception as e:
        st.session_state.sync_status = ("err",f"删除失败：{e}")
        return False

def sb_test():
    try:
        r = requests.get(f"{SB_URL}/rest/v1/{SB_TABLE}?select=id&limit=1",
                         headers=_hdrs(), timeout=8)
        r.raise_for_status()
        return True,"✅ 连接成功"
    except Exception as e:
        return False,f"❌ 连接失败：{e}"


# ════════════════════════════════════════════════
# 字段映射
# ════════════════════════════════════════════════
def local_to_cloud(r):
    return {
        "id":r["id"],"mold_id":r.get("moldId",""),"part_no":r.get("partNo",""),
        "part_ver":r.get("partVer",""),"type":r.get("type",""),
        "severity":r.get("severity","medium"),"date":r.get("date",""),
        "shift":r.get("shift",""),"machine_no":r.get("machineNo",""),
        "description":r.get("description",""),"reason":r.get("reason",""),
        "images":json.dumps(r.get("images",[]),ensure_ascii=False),
        "updated_at":datetime.utcnow().isoformat(),
    }

def cloud_to_local(r):
    imgs = []
    try: imgs = json.loads(r.get("images") or "[]")
    except: pass
    return {
        "id":r["id"],"moldId":r.get("mold_id",""),"partNo":r.get("part_no",""),
        "partVer":r.get("part_ver",""),"type":r.get("type",""),
        "severity":r.get("severity","medium"),"date":r.get("date",""),
        "shift":r.get("shift",""),"machineNo":r.get("machine_no",""),
        "description":r.get("description",""),"reason":r.get("reason",""),
        "images":imgs,"_synced":True,
    }


# ════════════════════════════════════════════════
# 同步
# ════════════════════════════════════════════════
def sync_from_supabase():
    st.session_state.sync_status = ("ing","同步中…")
    rows = sb_get_all()
    if st.session_state.sync_status[0] == "err":
        return
    cloud = [cloud_to_local(r) for r in rows]
    cloud_ids = {str(r["id"]) for r in cloud}
    pending = [r for r in st.session_state.data_store
               if not r.get("_synced") and str(r["id"]) not in cloud_ids]
    pushed = 0
    for rec in pending:
        if sb_upsert(rec):
            rec["_synced"] = True
            pushed += 1
    st.session_state.data_store = cloud + [r for r in pending if not r.get("_synced")]
    msg = f"已同步 {len(cloud)} 条"
    if pushed: msg += f"，补传 {pushed} 条"
    st.session_state.sync_status = ("ok", msg)
    st.session_state.synced_once = True


# ════════════════════════════════════════════════
# DeepSeek
# ════════════════════════════════════════════════
def ds_call(sys_p, user_p):
    try:
        r = requests.post(DS_API,
            headers={"Content-Type":"application/json","Authorization":f"Bearer {DEEPSEEK_KEY}"},
            json={"model":"deepseek-reasoner","max_tokens":8000,"temperature":1,
                  "messages":[{"role":"system","content":sys_p},{"role":"user","content":user_p}]},
            timeout=120)
        data = r.json()
        if "error" in data:
            st.error(f"DeepSeek 错误：{data['error']['message']}")
            return None
        return data["choices"][0]["message"]["content"]
    except Exception as e:
        st.error(f"AI 请求失败：{e}")
        return None


# ════════════════════════════════════════════════
# 工具函数
# ════════════════════════════════════════════════
def gen_file_no():
    n = datetime.now()
    return f"QCL-{str(n.year)[-2:]}{str(n.month).zfill(2)}-{str(len(st.session_state.saved_checklists)+1).zfill(3)}"

def sync_badge_html():
    s, msg = st.session_state.sync_status
    c  = {"ok":"#2e7d32","ing":"#bf360c","err":"#c62828"}
    bg = {"ok":"#e6f4ea","ing":"#fff4e5","err":"#fde8e8"}
    return (f'<div style="display:inline-flex;align-items:center;gap:6px;background:{bg[s]};'
            f'color:{c[s]};padding:5px 13px;border-radius:20px;font-size:12px;font-weight:600;'
            f'border:1px solid {c[s]}40;">● {msg}</div>')

def card_html(d):
    """生成记录卡片的纯 HTML，绝不包含任何可被 Streamlit 二次渲染的组件"""
    sev = d.get("severity","medium")
    fc, bg = SEV_COLOR.get(sev,("#333","#eee"))
    mold_id   = d.get("moldId","") or ""
    desc      = d.get("description","") or ""
    reason    = d.get("reason","") or ""
    rec_type  = d.get("type","") or ""
    rec_date  = d.get("date","") or ""
    machine   = d.get("machineNo","") or ""
    shift_v   = d.get("shift","") or ""
    shift_txt = SHIFT_LABEL.get(shift_v,"")
    synced    = d.get("_synced",False)
    synced_txt= "☁ 已同步" if synced else "💾 本地"
    synced_col= "#2e7d32" if synced else "#bf360c"
    imgs      = d.get("images",[]) or []

    extras = []
    if shift_txt and shift_txt != "—": extras.append(f"<span>{shift_txt}</span>")
    if machine: extras.append(f"<span>🔧 {machine}</span>")
    if imgs: extras.append(f"<span>🖼 {len(imgs)}张</span>")
    extras.append(f'<span style="color:{synced_col};font-weight:600;">{synced_txt}</span>')
    extra_html = "".join(extras)

    reason_html = ""
    if reason:
        reason_html = (f'<div style="font-size:11px;color:#7f8fa8;margin-top:4px;line-height:1.5;">'
                       f'<b>原因：</b>{reason}</div>')

    return f"""
<div style="background:#fff;border:1px solid #d4dae6;border-radius:10px;
  overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:2px;">
  <div style="background:{bg};padding:7px 12px;display:flex;
    justify-content:space-between;align-items:center;">
    <span style="color:{fc};font-weight:700;font-size:11px;">{SEV_LABEL.get(sev,sev)}</span>
    <span style="font-size:11px;color:#1565c0;font-weight:700;
      font-family:monospace;">{mold_id}</span>
  </div>
  <div style="padding:10px 12px;">
    <div style="font-size:13px;font-weight:700;color:#1e2a3b;margin-bottom:4px;">{rec_type}</div>
    <div style="font-size:12px;color:#445068;line-height:1.6;">{desc}</div>
    {reason_html}
    <div style="margin-top:8px;font-size:10px;color:#7f8fa8;
      display:flex;gap:10px;flex-wrap:wrap;align-items:center;">
      <span>📅 {rec_date}</span>
      {extra_html}
    </div>
  </div>
</div>"""


# ════════════════════════════════════════════════
# Excel 导出
# ════════════════════════════════════════════════
def export_excel(records):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "外观不良台账"
    headers = ["序号","模具号","零件号","版本","不良类型","严重程度",
               "发现日期","班次","机床号","现象描述","原因分析","同步状态"]
    widths  = [5,12,12,8,10,10,12,8,10,32,22,8]
    hf = PatternFill("solid",fgColor="1565C0"); bf = Font(color="FFFFFF",bold=True,size=10)
    thin = Side(style="thin",color="CCCCCC"); bdr = Border(left=thin,right=thin,top=thin,bottom=thin)
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        c=ws.cell(row=1,column=ci,value=h); c.fill=hf; c.font=bf; c.border=bdr
        c.alignment=Alignment(horizontal="center",vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[1].height=22
    for ri,d in enumerate(records,2):
        row=[ri-1,d.get("moldId",""),d.get("partNo",""),d.get("partVer",""),
             d.get("type",""),SEV_LABEL.get(d.get("severity",""),d.get("severity","")),
             d.get("date",""),
             "白班" if d.get("shift")=="day" else "夜班" if d.get("shift")=="night" else "",
             d.get("machineNo",""),d.get("description",""),d.get("reason",""),
             "已同步" if d.get("_synced") else "本地"]
        for ci,val in enumerate(row,1):
            c=ws.cell(row=ri,column=ci,value=val); c.border=bdr
            c.alignment=Alignment(vertical="center",wrap_text=(ci in(10,11)))
        if ri%2==0:
            for ci in range(1,len(headers)+1):
                ws.cell(row=ri,column=ci).fill=PatternFill("solid",fgColor="F0F4FA")
    ws.freeze_panes="A2"
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

def export_checklist_excel(cl):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="缺陷清单"
    hf=PatternFill("solid",fgColor="1565C0"); bf=Font(color="FFFFFF",bold=True,size=10)
    thin=Side(style="thin",color="CCCCCC"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    for row in [["文件编号",cl.get("fileNo",""),"","文件名称",cl.get("title","")],
                ["创建日期",cl.get("date",""),"","版本",cl.get("revision","A")],[]]:
        ws.append(row)
    cols=["序号","检查项目","模具/零件号","不良类型","严重程度","判定标准","频次","检查方法","处置方法","备注"]
    widths=[5,26,16,12,10,26,8,16,26,14]; hr=ws.max_row+1
    for ci,(h,w) in enumerate(zip(cols,widths),1):
        c=ws.cell(row=hr,column=ci,value=h); c.fill=hf; c.font=bf; c.border=bdr
        c.alignment=Alignment(horizontal="center",vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width=w
    for item in cl.get("items",[]):
        ri=ws.max_row+1
        for ci,val in enumerate([item.get("no",""),item.get("item",""),item.get("mold",""),
            item.get("type",""),item.get("severity",""),item.get("standard",""),
            item.get("frequency",""),item.get("method",""),item.get("action",""),item.get("remark","")],1):
            c=ws.cell(row=ri,column=ci,value=val); c.border=bdr
            c.alignment=Alignment(vertical="center",wrap_text=True)
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# ════════════════════════════════════════════════
# 侧边栏
# ════════════════════════════════════════════════
def render_sidebar():
    with st.sidebar:
        st.markdown("""
<div style="background:linear-gradient(135deg,#1565c0,#00695c);padding:11px 14px;
  border-radius:10px;margin-bottom:12px;">
  <div style="color:#fff;font-size:14px;font-weight:700;">🏭 外观不良管理 V3</div>
  <div style="color:rgba(255,255,255,.65);font-size:10px;font-family:monospace;margin-top:2px;">
    外观缺陷智能管理系统 · PY版</div>
</div>""", unsafe_allow_html=True)

        st.markdown(sync_badge_html(), unsafe_allow_html=True)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        edit = st.session_state.edit_record
        st.markdown("#### 📝 数据录入")

        mold_names = [m["name"] for m in st.session_state.mold_list]
        mold_display = ["— 请选择模具 —"] + [
            f"{m['name']} ｜ 零件:{m.get('partNo') or '-'}"
            for m in st.session_state.mold_list
        ]
        mold_map = {label: mold for label, mold in zip(mold_display[1:], st.session_state.mold_list)}
        mold_idx = 0
        if edit and edit.get("moldId") in mold_names:
            mold_idx = mold_names.index(edit["moldId"]) + 1

        sel_mold_str = st.selectbox("模具编号 *", mold_display, index=mold_idx, key="sb_mold")
        mold_obj = mold_map.get(sel_mold_str)
        if mold_obj:
            pn = mold_obj.get("partNo","") or "—"
            pv = mold_obj.get("partVer","") or "—"
            pm = mold_obj.get("productName","") or ""
            st.caption(f"零件号：**{pn}** ｜ 版本：**{pv}**" + (f" ｜ {pm}" if pm else ""))

        c1,c2 = st.columns(2)
        with c1:
            default_date = date.today()
            if edit and edit.get("date"):
                try: default_date = datetime.strptime(edit["date"],"%Y-%m-%d").date()
                except: pass
            rec_date = st.date_input("日期 *", value=default_date, key="sb_date")
            sev_keys = list(SEV_LABEL.keys())
            sev_idx = sev_keys.index(edit.get("severity","medium")) if edit else 1
            severity = st.selectbox("严重程度 *", sev_keys,
                format_func=lambda x: SEV_LABEL[x], index=sev_idx, key="sb_sev")
        with c2:
            shift_keys = ["","day","night"]
            shift_idx = shift_keys.index(edit.get("shift","")) if edit and edit.get("shift","") in shift_keys else 0
            shift = st.selectbox("班次", shift_keys,
                format_func=lambda x:{"":"— 请选择 —","day":"☀ 白班","night":"🌙 夜班"}.get(x,x),
                index=shift_idx, key="sb_shift")
            types = st.session_state.defect_types
            type_idx = types.index(edit["type"]) if edit and edit.get("type") in types else 0
            defect_type = st.selectbox("不良类型 *", types, index=type_idx, key="sb_type")

        machine_no  = st.text_input("机床号", value=edit.get("machineNo","") if edit else "", key="sb_machine")
        description = st.text_area("现象描述 *", value=edit.get("description","") if edit else "", height=75, key="sb_desc")
        reason      = st.text_area("原因分析", value=edit.get("reason","") if edit else "", height=60, key="sb_reason")
        imgs_up     = st.file_uploader("上传图片", type=["jpg","jpeg","png","webp"],
                                        accept_multiple_files=True, key="sb_imgs")

        b1,b2 = st.columns(2)
        with b1:
            submit = st.button("✔ 提交记录" if not edit else "✔ 保存修改",
                               use_container_width=True, type="primary", key="btn_submit")
        with b2:
            cancel = st.button("✖ 取消编辑", use_container_width=True, key="btn_cancel")

        if cancel:
            st.session_state.edit_record = None
            st.rerun()

        if submit:
            if sel_mold_str == "— 请选择模具 —" and not edit:
                st.error("请先选择模具号")
            elif not description.strip():
                st.error("请填写现象描述")
            else:
                mold_id  = mold_obj["id"] if mold_obj else (edit.get("moldId","") if edit else "")
                part_no  = mold_obj.get("partNo","") if mold_obj else (edit.get("partNo","") if edit else "")
                part_ver = mold_obj.get("partVer","") if mold_obj else (edit.get("partVer","") if edit else "")
                images   = list(edit.get("images",[]) if edit else [])
                if imgs_up:
                    for f in imgs_up:
                        b64 = base64.b64encode(f.read()).decode()
                        images.append(f"data:{f.type or 'image/jpeg'};base64,{b64}")
                record = {
                    "id":          edit["id"] if edit else int(time.time()*1000),
                    "moldId":      mold_id,
                    "partNo":      part_no,
                    "partVer":     part_ver,
                    "type":        defect_type,
                    "severity":    severity,
                    "date":        str(rec_date),
                    "shift":       shift,
                    "machineNo":   machine_no.strip(),
                    "description": description.strip(),
                    "reason":      reason.strip(),
                    "images":      images,
                    "_synced":     False,
                }
                with st.spinner("提交中…"):
                    ok = sb_upsert(record)
                if ok:
                    record["_synced"] = True
                    st.session_state.sync_status = ("ok","已同步")
                ds = st.session_state.data_store
                if edit:
                    idx = next((i for i,r in enumerate(ds) if r["id"]==edit["id"]), None)
                    if idx is not None: ds[idx] = record
                else:
                    ds.insert(0, record)
                st.session_state.edit_record = None
                st.success("✅ 记录已保存")
                st.rerun()

        st.divider()
        st.caption("基础资料管理（模具/不良类型）已移到右侧「不良台账」页面。")

        with st.expander("☁ Supabase 连接测试"):
            if st.button("🔌 测试连接", use_container_width=True):
                ok,msg = sb_test()
                st.success(msg) if ok else st.error(msg)


# ════════════════════════════════════════════════
# 主区：基础资料管理（横向）
# ════════════════════════════════════════════════
def render_master_data_panel():
    st.markdown("##### ⚙ 基础资料管理")
    lc, rc = st.columns(2)

    with lc:
        st.markdown("###### 🔧 模具管理")
        with st.form("main_form_add_mold", clear_on_submit=True):
            m1,m2,m3,m4,m5 = st.columns([2.2,1.6,1.2,2.2,0.9])
            with m1:
                m_name = st.text_input("模具号*", placeholder="如 1507#")
            with m2:
                m_part = st.text_input("零件号", placeholder="零件号")
            with m3:
                m_ver = st.text_input("版本", placeholder="版本")
            with m4:
                m_pname = st.text_input("产品名称", placeholder="如 刹车卡钳支架")
            with m5:
                add_m = st.form_submit_button("添加", use_container_width=True)

        if add_m:
            mn = normalize_mold_code(m_name)
            existing = {normalize_mold_code(m.get("id") or m.get("name")) for m in st.session_state.mold_list}
            if not mn:
                st.error("请输入模具号")
            elif mn in existing:
                st.error("该模具号已存在")
            else:
                st.session_state.mold_list.append({
                    "id":mn, "name":mn, "partNo":m_part.strip(),
                    "partVer":m_ver.strip(), "productName":m_pname.strip()
                })
                st.success(f"已添加模具：{mn}")
                st.rerun()

        if st.session_state.mold_list:
            h1,h2,h3,h4,h5 = st.columns([1.6,1.6,1.0,2.0,0.9])
            h1.markdown("**模具号**")
            h2.markdown("**零件号**")
            h3.markdown("**版本**")
            h4.markdown("**产品名称**")
            h5.markdown("**操作**")
            for i, m in enumerate(st.session_state.mold_list):
                c1,c2,c3,c4,c5 = st.columns([1.6,1.6,1.0,2.0,0.9])
                c1.write(m["name"])
                c2.write(m.get("partNo","") or "-")
                c3.write(m.get("partVer","") or "-")
                c4.write(m.get("productName","") or "-")
                if c5.button("删除", key=f"main_dlm_{m['id']}_{i}", use_container_width=True):
                    st.session_state.mold_list = [x for x in st.session_state.mold_list if x["id"] != m["id"]]
                    st.rerun()

    with rc:
        st.markdown("###### 🏷 不良类型管理")
        with st.form("main_form_add_type", clear_on_submit=True):
            t1,t2 = st.columns([5,1])
            with t1:
                new_type = st.text_input("新增不良类型", placeholder="如：夹渣")
            with t2:
                add_t = st.form_submit_button("添加", use_container_width=True)

        if add_t:
            nt = new_type.strip()
            if not nt:
                st.error("请输入不良类型")
            elif nt in st.session_state.defect_types:
                st.error("该类型已存在")
            else:
                st.session_state.defect_types.append(nt)
                st.success(f"已添加类型：{nt}")
                st.rerun()

        if st.session_state.defect_types:
            h1,h2 = st.columns([5,1])
            h1.markdown("**不良类型**")
            h2.markdown("**操作**")
            for i, t in enumerate(st.session_state.defect_types):
                tc1,tc2 = st.columns([5,1])
                tc1.write(t)
                if tc2.button("删除", key=f"main_dlt_{i}", use_container_width=True):
                    st.session_state.defect_types.pop(i)
                    st.rerun()


# ════════════════════════════════════════════════
# 台账 Tab
# ════════════════════════════════════════════════
def render_records_tab():
    render_master_data_panel()
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    fc = st.columns([2,1.5,1.5,1.5,1.5,1])
    mold_opts = ["全部模具"] + [m["name"] for m in st.session_state.mold_list]
    with fc[0]: f_mold  = st.selectbox("模具", mold_opts, label_visibility="collapsed", key="f_mold")
    with fc[1]:
        f_sev = st.selectbox("严重程度", ["all"]+list(SEV_LABEL.keys()),
            format_func=lambda x:"全部严重程度" if x=="all" else SEV_LABEL.get(x,x),
            label_visibility="collapsed", key="f_sev")
    with fc[2]:
        f_shift = st.selectbox("班次", ["all","day","night"],
            format_func=lambda x:{"all":"全部班次","day":"白班","night":"夜班"}.get(x,x),
            label_visibility="collapsed", key="f_shift")
    with fc[3]: f_from = st.date_input("开始日期", value=None, label_visibility="collapsed", key="f_from")
    with fc[4]: f_to   = st.date_input("结束日期",  value=None, label_visibility="collapsed", key="f_to")
    with fc[5]:
        if st.button("清空筛选", use_container_width=True, key="btn_clr"):
            for k in ["f_mold","f_sev","f_shift","f_from","f_to"]:
                if k in st.session_state: del st.session_state[k]
            st.rerun()

    ds = st.session_state.data_store
    filtered = [d for d in ds if
        (f_mold=="全部模具" or d.get("moldId","")==f_mold) and
        (f_sev=="all"      or d.get("severity")==f_sev) and
        (f_shift=="all"    or d.get("shift")==f_shift) and
        (not f_from or d.get("date","")>=str(f_from)) and
        (not f_to   or d.get("date","")<=str(f_to))
    ]

    by_sev = {k:sum(1 for d in filtered if d.get("severity")==k) for k in SEV_LABEL}
    m1,m2,m3,m4 = st.columns(4)
    m1.metric("📋 总记录", len(filtered))
    m2.metric("🔴 严重/致命", by_sev["high"]+by_sev["critical"])
    m3.metric("🟠 一般",      by_sev["medium"])
    m4.metric("🟢 轻微",      by_sev["low"])

    op = st.columns([1.2,1.2,5])
    with op[0]:
        if st.button("🔄 同步云端", use_container_width=True, key="btn_sync"):
            with st.spinner("同步中…"):
                sync_from_supabase()
            st.rerun()
    with op[1]:
        if filtered:
            st.download_button("📥 导出Excel", data=export_excel(filtered),
                file_name=f"外观不良台账_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="btn_exp")

    st.markdown("---")

    if not filtered:
        st.markdown("""
<div style="text-align:center;padding:48px;color:#7f8fa8;">
  <div style="font-size:36px;margin-bottom:12px;">📭</div>
  <div>暂无记录，请在左侧录入数据</div>
</div>""", unsafe_allow_html=True)
        return

    cols_n = 3
    for i in range(0, len(filtered), cols_n):
        row_items = filtered[i:i+cols_n]
        cols = st.columns(cols_n)
        for col, d in zip(cols, row_items):
            with col:
                # ★ 纯 HTML 渲染，避免源码暴露
                st.markdown(card_html(d), unsafe_allow_html=True)

                imgs = d.get("images",[]) or []
                if imgs:
                    with st.expander(f"🖼 查看图片（{len(imgs)}张）"):
                        for src in imgs[:6]:
                            st.image(src, use_column_width=True)

                bc1,bc2 = st.columns(2)
                with bc1:
                    if st.button("✏ 编辑", key=f"edit_{d['id']}", use_container_width=True):
                        st.session_state.edit_record = d
                        st.rerun()
                with bc2:
                    if st.button("🗑 删除", key=f"del_{d['id']}", use_container_width=True):
                        st.session_state.data_store = [r for r in st.session_state.data_store if r["id"]!=d["id"]]
                        with st.spinner("删除中…"):
                            sb_delete(d["id"])
                        st.session_state.sync_status = ("ok","删除已同步")
                        st.rerun()
                st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)


# ════════════════════════════════════════════════
# 缺陷清单 Tab
# ════════════════════════════════════════════════
def render_checklist_tab():
    st.markdown("##### 🤖 AI 缺陷检查清单")

    cb1,cb2 = st.columns(2)
    with cb1:
        gen_btn = st.button("🤖 AI 生成缺陷清单", type="primary", use_container_width=True)
    with cb2:
        if st.session_state.saved_checklists:
            cl0 = st.session_state.saved_checklists[0]
            st.download_button("📥 导出Excel清单",
                data=export_checklist_excel(cl0),
                file_name=f"缺陷清单_{cl0.get('fileNo','')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

    if gen_btn:
        ds = st.session_state.data_store
        summary = "\n".join([
            f"模具:{d.get('moldId','')} 零件:{d.get('partNo','-')} 类型:{d.get('type','')} "
            f"程度:{SEV_LABEL.get(d.get('severity',''),'')} 描述:{d.get('description','')}"
            for d in ds[:60]
        ])
        mold_info = "、".join([
            f"{m['name']}(零件:{m.get('partNo','-')} 版本:{m.get('partVer','-')})"
            for m in st.session_state.mold_list
        ])
        fn = gen_file_no(); today = datetime.now().strftime("%Y/%m/%d")
        with st.spinner("DeepSeek-R1 分析中，请稍候（约30秒）…"):
            result = ds_call(
                "你是资深精密铸造质量工程师，熟悉IATF16949。请严格只输出合法JSON，不加任何说明文字或markdown符号。",
                f"根据以下不良记录，生成正式缺陷检查清单：\n模具：{mold_info}\n记录：\n{summary}\n\n"
                f'输出JSON格式：{{"title":"外观不良检查清单","fileNo":"{fn}","date":"{today}",'
                f'"revision":"A","scope":"适用范围","items":[{{"no":1,"item":"检查项目","mold":"模具号",'
                f'"type":"类型","severity":"程度","standard":"判定标准","frequency":"频次",'
                f'"method":"方法","action":"处置","remark":"备注"}}],"notes":["注意事项"]}}',
            )
        if result:
            try:
                parsed = json.loads(result.replace("```json","").replace("```","").strip())
                st.session_state.saved_checklists.insert(0, parsed)
                st.success("✅ 清单已生成")
                st.rerun()
            except:
                st.code(result, language="json")

    st.divider()

    if st.session_state.saved_checklists:
        cl = st.session_state.saved_checklists[0]
        st.markdown(f"""
<div style="background:linear-gradient(135deg,#f0f7ff,#e8f4ff);border:1px solid #b8d9f8;
  border-radius:10px;padding:12px 16px;margin-bottom:12px;
  display:flex;justify-content:space-between;align-items:flex-start;">
  <div>
    <div style="font-size:15px;font-weight:700;color:#1e2a3b;">{cl.get('title','缺陷检查清单')}</div>
    <div style="font-size:11px;color:#445068;margin-top:3px;">{cl.get('scope','')}</div>
  </div>
  <div style="text-align:right;">
    <div style="font-family:monospace;color:#bf360c;font-size:11px;font-weight:700;">{cl.get('fileNo','')}</div>
    <div style="font-size:10px;color:#7f8fa8;">Rev.{cl.get('revision','A')} · {cl.get('date','')}</div>
  </div>
</div>""", unsafe_allow_html=True)

        items = cl.get("items",[])
        if items:
            import pandas as pd
            df = pd.DataFrame([{
                "序号":it.get("no",""),"检查项目":it.get("item",""),
                "模具/零件":it.get("mold","-"),"类型":it.get("type","-"),
                "程度":it.get("severity","-"),"判定标准":it.get("standard","-"),
                "频次":it.get("frequency","-"),"方法":it.get("method","-"),
                "处置":it.get("action","-"),"备注":it.get("remark",""),
            } for it in items])
            st.dataframe(df, use_container_width=True, hide_index=True, height=380)

        notes = cl.get("notes",[])
        if notes:
            st.markdown("**注意事项：**")
            for i,n in enumerate(notes,1): st.markdown(f"{i}. {n}")
    else:
        st.info("暂无清单，点击上方「AI 生成缺陷清单」")

    st.divider()
    with st.expander("📄 上传缺陷清单模板（Excel / CSV / TXT）"):
        up = st.file_uploader("选择文件", type=["xlsx","xls","csv","txt"], key="tpl_up")
        if up and st.button("🤖 AI 解析模板", key="btn_parse"):
            if up.name.endswith((".xlsx",".xls")):
                wb2 = openpyxl.load_workbook(io.BytesIO(up.read()))
                ws2 = wb2.active
                content = "\n".join([",".join([str(c.value or "") for c in row]) for row in ws2.iter_rows()])
            else:
                content = up.read().decode("utf-8",errors="ignore")
            fn = gen_file_no(); today = datetime.now().strftime("%Y/%m/%d")
            with st.spinner("DeepSeek-R1 解析中…"):
                result = ds_call(
                    "你是质量工程师。将模板整理为标准缺陷清单JSON，严格只输出合法JSON，不加任何说明。",
                    f'模板内容：\n{content[:4000]}\n\n输出：{{"title":"...","fileNo":"{fn}",'
                    f'"date":"{today}","revision":"A","scope":"...","items":[{{"no":1,"item":"...",'
                    f'"mold":"...","type":"...","severity":"...","standard":"...","frequency":"...",'
                    f'"method":"...","action":"...","remark":"..."}}],"notes":[]}}',
                )
            if result:
                try:
                    parsed = json.loads(result.replace("```json","").replace("```","").strip())
                    st.session_state.saved_checklists.insert(0, parsed)
                    st.success("✅ 解析成功"); st.rerun()
                except: st.code(result)


# ════════════════════════════════════════════════
# 报告 Tab
# ════════════════════════════════════════════════
def render_report_tab():
    st.markdown("##### 📊 AI 质量分析报告")
    rc1,rc2 = st.columns([2,4])
    with rc1:
        rpt_type = st.selectbox("报告周期", ["本周","本月","自定义"], key="rpt_type")
    from_date = to_date = None
    if rpt_type == "自定义":
        d1,d2 = st.columns(2)
        with d1: from_date = st.date_input("开始日期", key="rpt_from")
        with d2: to_date   = st.date_input("结束日期",  key="rpt_to")

    if st.button("🤖 AI 生成报告", type="primary"):
        now = datetime.now()
        if rpt_type == "本周":
            from_date = date(now.year, now.month, now.day - now.weekday())
            to_date   = date.today()
        elif rpt_type == "本月":
            from_date = date(now.year, now.month, 1)
            to_date   = date.today()
        if not from_date or not to_date:
            st.warning("请选择日期范围")
        else:
            ds = st.session_state.data_store
            data = [d for d in ds if str(from_date)<=d.get("date","")<=str(to_date)]
            by_sev = {k:sum(1 for d in data if d.get("severity")==k) for k in SEV_LABEL}
            summary = "\n".join([
                f"[{d.get('date','')}]{'白班' if d.get('shift')=='day' else '夜班' if d.get('shift')=='night' else ''} "
                f"模具:{d.get('moldId','')} 类型:{d.get('type','')} 程度:{SEV_LABEL.get(d.get('severity',''),'')} "
                f"机床:{d.get('machineNo','-')} 描述:{d.get('description','')}"
                f"{'  分析:'+d.get('reason') if d.get('reason') else ''}"
                for d in data
            ])
            with st.spinner("DeepSeek-R1 撰写报告，请稍候（约30-60秒）…"):
                result = ds_call(
                    "你是精密铸造工厂质量经理，负责撰写专业质量周报/月报用于管理层会议汇报。报告需要数据分析、趋势研判、风险预警和改善建议，语言简洁专业，结构清晰。",
                    f"请根据 {from_date} 至 {to_date} 的外观不良数据生成质量分析报告：\n\n"
                    f"统计摘要：总计{len(data)}条 | 轻微{by_sev['low']} | 一般{by_sev['medium']} | "
                    f"严重{by_sev['high']} | 致命{by_sev['critical']}\n\n"
                    f"详细记录：\n{summary or '（该时段无记录）'}\n\n"
                    f"报告要求：①执行摘要 ②数据概览（表格） ③主要不良分析 ④趋势与风险 ⑤改善建议 ⑥重点关注项",
                )
            if result:
                st.session_state.last_report = result
                st.session_state.last_report_range = f"{from_date}至{to_date}"

    if st.session_state.last_report:
        rpt = st.session_state.last_report
        st.text_area("📋 报告内容（可复制）", value=rpt, height=520, key="rpt_out")
        st.download_button("⬇ 导出 TXT 文件", data=rpt.encode("utf-8"),
            file_name=f"质量报告_{st.session_state.last_report_range}.txt",
            mime="text/plain")
    else:
        st.info("选择报告周期后点击「AI 生成报告」")


# ════════════════════════════════════════════════
# 主程序
# ════════════════════════════════════════════════
def main():
    init_state()

    # 启动时自动同步
    if not st.session_state.synced_once:
        with st.spinner("正在连接 Supabase 并同步数据…"):
            sync_from_supabase()

    render_sidebar()

    st.markdown("""
<div style="display:flex;align-items:center;margin-bottom:10px;
  padding:10px 14px;border:1px solid #d5e1ee;border-radius:12px;
  background:linear-gradient(135deg,#ffffff,#f3f8fd);">
  <h2 style="margin:0;font-size:24px;font-weight:900;color:#183149;">
    🏭 外观不良智能管理系统
  </h2>
</div>""", unsafe_allow_html=True)

    tab1,tab2,tab3 = st.tabs(["📋 不良台账","📄 AI缺陷清单","📊 AI质量报告"])
    with tab1: render_records_tab()
    with tab2: render_checklist_tab()
    with tab3: render_report_tab()


if __name__ == "__main__":
    main()
